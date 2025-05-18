# import streamlit as st
# import pandas as pd
# import sqlite3
# import re

# # Title of the app
# st.header("Management Information System")

# # Sidebar: File upload and sheet selection
# with st.sidebar:
#     file = st.file_uploader("Upload an Excel file", type=["xlsx", "xls"])

# if file is not None:
#     try:
#         # Read all sheets from the uploaded Excel file without headers
#         all_sheets = pd.read_excel(file, sheet_name=None, header=None)
#         sheet_names = list(all_sheets.keys())

#         # Sidebar: Let user select which sheet to work on
#         with st.sidebar:
#             selected_sheet = st.selectbox("Select Sheet", sheet_names)

#         # Load the selected sheet into a DataFrame
#         df = all_sheets[selected_sheet]

#         # Mapping keywords to classification categories
#         classification_map = {
#             "advance to vendor": "Advance Given",
#             "sundry debtors": "Debtors",
#             "sundry creditors": "Creditors",
#             "retention money": "Retention",
#             "secured loan": "Secured",
#             "unsecured loan": "Unsecured",
#             "od limit": "OD",
#             "project wise billing": "sales"
#         }

#         # Flatten DataFrame rows to lowercase strings for keyword detection
#         flat_text = df.astype(str).apply(lambda x: ' '.join(x.dropna()), axis=1).str.lower()

#         detected_classification = None
#         # Check for classification keywords in data rows
#         for keyword, classification in classification_map.items():
#             pattern = rf"\b{re.escape(keyword)}\b"  # Word boundary for exact matching
#             if flat_text.str.contains(pattern, regex=True).any():
#                 detected_classification = classification
#                 st.info(f"🟢 '{keyword.title()}' found in data — setting classification: **{classification}**")
#                 break

#         # Regex pattern to detect a date (various formats)
#         date_pattern = r"\b(?:\d{1,2}[-/thstndrd\s\.])?(?:\d{1,2}[-/\s\.])?(?:\d{2,4})\b"
#         report_date = None
#         # Search first 10 rows for any date-like string
#         for i in range(min(10, len(df))):
#             row = df.iloc[i].dropna().astype(str)
#             for val in row:
#                 match = re.search(date_pattern, val)
#                 if match:
#                     parsed_date = pd.to_datetime(match.group(), errors='coerce')
#                     if pd.notnull(parsed_date):
#                         report_date = parsed_date.date()
#                         break
#             if report_date:
#                 break

#         if report_date:
#             st.write("📅 Report Date Detected:", report_date)

#         # Keywords to detect header row
#         keywords = ["particulars", "party name", "lender's name", "item head"]
#         target_index = None
#         matched_keyword = None

#         # Find header row by checking if any keyword is present in the row
#         for index, row in df.iterrows():
#             # Convert row to a clean lowercase string without slashes or dashes
#             row_values_str = ' '.join(
#                 ''.join(map(str, row.values))
#                 .lower()
#                 .replace('/', ' ')
#                 .replace('-', ' ')
#                 .split()
#             )
#             for kw in keywords:
#                 if kw in row_values_str:
#                     target_index = index
#                     matched_keyword = kw
#                     break
#             if target_index is not None:
#                 break

#         # If header found, set it as column headers and drop all rows above it
#         if target_index is not None:
#             st.write(f"✅ Found header using keyword: `{matched_keyword}` at row {target_index}")
#             df.columns = df.iloc[target_index]
#             df = df.iloc[target_index + 1:].reset_index(drop=True)
#         else:
#             st.warning("No header row found with keywords: 'particulars', 'party name', or 'lender's name'")

#         # Drop columns with all NaN headers or all NaN values
#         df = df.loc[:, df.columns.notna()]
#         df = df.dropna(axis=1, how='all')

#         # Remove any rows containing "total" (case-insensitive) in any cell
#         df = df[~df.apply(lambda row: row.astype(str).str.contains("total", case=False, na=False)).any(axis=1)]

#         # Drop a 'Total' column if it exists
#         if "Total" in df.columns:
#             df = df.drop("Total", axis=1)

#         # Fix common column header typos
#         if "INstallment No" in df.columns:
#             df = df.rename(columns={"INstallment No": "Installment No"})
#         if "Rate of Interest" in df.columns:
#             df = df.rename(columns={"Rate of Interest": "Interest"})
#             # Add a serial number column starting at 1
#             df.insert(0, 'S. No.', range(1, len(df) + 1))

#         # If sales classification, reduce dataframe to specific columns
#         if "Project / Item head" in df.columns:
#             df = df.iloc[:, [0, 1, -1]]
#             df = df.rename(columns={df.columns[-1]: "Value"})

#         st.subheader("Final Data")

#         # Add extracted metadata columns for Date and Classification
#         df['Date'] = report_date
#         df['Classification'] = detected_classification

#         # Drop rows with more than 4 null values (likely empty rows)
#         df = df[df.isnull().sum(axis=1) <= 4]

#         flag = -1  # Flag to indicate classification processing type

#         # Clean and process data based on detected classification
#         if detected_classification in ["Advance Given", "Debtors", "Creditors", "Retention"]:
#             if "Particulars" in df.columns:
#                 df = df.rename(columns={"Particulars": "Party Name"})
#             df = df[df["Party Name"].notnull()]  # Drop rows where Party Name is null
#             flag = 1
#             # Remove parentheses from column headers for consistency
#             df.columns = df.columns.str.replace("(", "", regex=False).str.replace(")", "", regex=False)

#         elif detected_classification in ["Secured", "Unsecured", "OD"]:
#             flag = 2

#         elif detected_classification == "sales":
#             flag = 3
#             # Drop rows where "Project / Item head" is null
#             df = df[df["Project / Item head"].notnull()]

#         # Display final cleaned DataFrame
#         st.write(df)

#         # Connect to SQLite database
#         conn = sqlite3.connect('dummy_wog.db')
#         cursor = conn.cursor()

#         # Button to push data into the SQL database
#         if st.button("Push to SQL"):
#             if flag == 1:
#                 # Insert data into CurrentAsset_CurrentLiabilities table
#                 for _, row in df.iterrows():
#                     cursor.execute("""
#                         INSERT INTO CurrentAsset_CurrentLiabilities (
#                             "Serial Number", 
#                             "Particulars", 
#                             "Less Than 120 Days", 
#                             "More Than 120 Days", 
#                             "Date", 
#                             "Classification"
#                         ) VALUES (?, ?, ?, ?, ?, ?)
#                     """, (
#                         row[0],  # S. No.
#                         row[1],  # Particulars
#                         row[2],  # Less Than 120 Days
#                         row[3],  # More Than 120 Days
#                         row[4],  # Date
#                         row[5]   # Classification
#                     ))

#             elif flag == 2:
#                 # Handle loans insertion based on classification type
#                 if detected_classification == "Unsecured":
#                     for _, row in df.iterrows():
#                         cursor.execute("""
#                             INSERT INTO Loans (
#                                 "Serial Number","LENDER'S NAME ", "Loan Amount", "Interest", "Period", "Emi Amount" ,"Emi Due Date","Installment No","Principle Remaining","Principle Outstanding","Date","Classification"
#                             ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
#                         """, (
#                             row['S. No.'],
#                             row["LENDER'S NAME "],
#                             row['Loan amount'],
#                             row['Interest'],
#                             row['Period (months)'],
#                             row['EMI AMOUNT'],
#                             row['EMI DUE DATE'],
#                             row['Installment No'],
#                             row['Principle Remaining'],
#                             row['Principle Outstanding'],
#                             row['Date'],
#                             row['Classification']
#                         ))
#                 elif detected_classification == "OD":
#                     for _, row in df.iterrows():
#                         cursor.execute("""
#                             INSERT INTO Loans (
#                                 "Serial Number", "Interest", "LENDER'S NAME ","Utilisation Amount","Balance Unused","Classification"
#                             ) VALUES (?, ?, ?, ?, ?, ?)
#                         """, (
#                             row['S. No.'],
#                             row['Interest'],
#                             row["LENDER'S NAME "],
#                             row['Utilisation Amount'],
#                             row['Balance Unused'],
#                             row['Classification']
#                         ))
#                 elif detected_classification == "Secured":
#                     for _, row in df.iterrows():
#                         cursor.execute("""
#                             INSERT INTO Loans (
#                                 "Serial Number","LENDER'S NAME ", "Loan Amount", "Interest", "Period", "Emi Amount" ,"Installment No","Principle Remaining","Principle Outstanding","Date","Classification"
#                             ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
#                         """, (
#                             row['S. No.'],
#                             row["LENDER'S NAME "],
#                             row['Loan amount'],
#                             row['Interest'],
#                             row['Period (months)'],
#                             row['EMI AMOUNT'],
#                             row['Installment No'],
#                             row['Principle Remaining'],
#                             row['Principle Outstanding'],
#                             row['Date'],
#                             row['Classification']
#                         ))

#             elif flag == 3:
#                 # Insert sales data into Sales table
#                 for _, row in df.iterrows():
#                     cursor.execute("""
#                         INSERT INTO Sales (
#                             "Serial Number", 
#                             "Item Head", 
#                             "Value", 
#                             "Date", 
#                             "Classification"
#                         ) VALUES (?, ?, ?, ?, ?)
#                     """, (
#                         row['S.No'],
#                         row['Project / Item head'],
#                         row['Value'],
#                         row['Date'],
#                         row['Classification']
#                     ))

#             # Commit changes and close connection
#             conn.commit()
#             conn.close()
#             st.success("Data saved to SQLite database successfully!")

#     except Exception as e:
#         # Handle any errors during processing
#         st.warning("Something went wrong")
#         st.text(f"Error: {e}")

# else:
#     # Prompt user to upload file if none provided
#     st.info("Please upload an Excel file to get started.")



#### using the workbook

 ## it  is in openxyl load_workbook
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import sqlite3
import re

st.header("Management Information System")

with st.sidebar:
    file = st.file_uploader("Upload an Excel file", type=["xlsx","xls"])

flag = -1

if file is not None:
    try:
        # Load workbook using openpyxl
        wb = load_workbook(filename=file, data_only=True)  # data_only=True to get values, not formulas
        sheetnames = wb.sheetnames
        
        with st.sidebar:
            selected_sheet = st.selectbox("Select Sheet", sheetnames)
        
        ws = wb[selected_sheet]

        # Read sheet data as generator of rows
        data = ws.values
        # Convert to list of rows (tuples)
        data = list(data)

        # Convert to DataFrame, assuming no header yet (like header=None in your original)
        df = pd.DataFrame(data)

        # The rest of your original processing logic starts here:
        classification_map = {
            "advance to vendor": "Advance Given",
            "sundry debtors": "Debtors",
            "sundry creditors": "Creditors",
            "retention money": "Retention",
            "secured loan": "Secured",
            "unsecured loan": "Unsecured",
            "od limit": "OD",
            "project wise billing": "sales"
        }

        flat_text = df.astype(str).apply(lambda x: ' '.join(x.dropna()), axis=1).str.lower()

        detected_classification = None
        for keyword, classification in classification_map.items():
            pattern = rf"\b{re.escape(keyword)}\b"
            if flat_text.str.contains(pattern, regex=True).any():
                detected_classification = classification
                st.info(f"🟢 '{keyword.title()}' found in data — setting classification: **{classification}**")
                break

        # Extract report date from the first few rows
        date_pattern = r"\b(?:\d{1,2}[-/thstndrd\s\.])?(?:\d{1,2}[-/\s\.])?(?:\d{2,4})\b"
        report_date = None
        for i in range(min(10, len(df))):
            row = df.iloc[i].dropna().astype(str)
            for val in row:
                match = re.search(date_pattern, val)
                if match:
                    try:
                        parsed_date = pd.to_datetime(match.group(), errors='coerce')
                        if pd.notnull(parsed_date):
                            report_date = parsed_date.date()
                            break
                    except:
                        continue
            if report_date:
                break

        if report_date:
            st.write("📅 Report Date Detected:", report_date)

        # Find header row by keywords
        keywords = ["particulars", "party name", "lender's name", "item head"]
        target_index = None
        matched_keyword = None

        for index, row1 in df.iterrows():
            row_values_str = ' '.join(
                ''.join(map(str, row1.values))
                .lower()
                .replace('/', ' ')
                .replace('-', ' ')
                .split()
            )

            for kw in keywords:
                if kw in row_values_str:
                    target_index = index
                    matched_keyword = kw
                    break
            if target_index is not None:
                break

        if target_index is not None:
            st.write(f"✅ Found header using keyword: `{matched_keyword}` at row {target_index}")
            df.columns = df.iloc[target_index]
            df = df.iloc[target_index + 1:].reset_index(drop=True)
        else:
            st.warning("No header row found with keywords: 'particulars', 'party name', or 'lender's name'")

        df = df.loc[:, df.columns.notna()]
        df = df.dropna(axis=1, how='all') 
        st.write(df.shape)

        # Remove rows with 'total'
        df = df[~df.apply(lambda row: row.astype(str).str.contains("total", case=False, na=False)).any(axis=1)]

        if "Total" in df.columns:
            df = df.drop("Total", axis=1)
        if "INstallment No" in df.columns:
            df = df.rename(columns={"INstallment No":"Installment No"})
        if "Rate of Interest" in df.columns:
            df = df.rename(columns={"Rate of Interest":"Interest"})
            df.insert(0, 'S. No.', range(1, len(df) + 1))

        if "Project / Item head" in df.columns:
            df = df.iloc[:, [0,1,-1]]
            df = df.rename(columns={df.columns[-1]: "Value"})

        st.subheader("Final Data")
        df['Date'] = report_date
        df['Classification'] = detected_classification

        df = df[df.isnull().sum(axis=1) <= 4]

        if detected_classification in ["Advance Given","Debtors","Creditors","Retention"]:
            if "Particulars" in df.columns:
                df = df.rename(columns={"Particulars":"Party Name"})
            df = df[df["Party Name"].notnull()]
            flag = 1
            df.columns = df.columns.str.replace("(", "", regex=False).str.replace(")", "", regex=False)

        elif detected_classification in ["Secured","Unsecured","OD"]:
            flag = 2
        elif detected_classification == "sales":
            flag = 3
            df = df[df["Project / Item head"].notnull()]

        st.write(df)

        # SQLite insertion part
        conn = sqlite3.connect('dummy_wog.db')
        cursor = conn.cursor()

        if st.button("Push to SQL"):
            if flag == 1:
                for _, row in df.iterrows():
                    cursor.execute("""
                        INSERT INTO CurrentAsset_CurrentLiabilities (
                            "Serial Number", 
                            "Particulars", 
                            "Less Than 120 Days", 
                            "More Than 120 Days", 
                            "Date", 
                            "Classification"
                        ) VALUES (?, ?, ?, ?, ?, ?)
                    """, (
                        row[0],
                        row[1],
                        row[2],
                        row[3],
                        row[4],
                        row[5]
                    ))

            if flag == 2:
                if detected_classification == "Unsecured":
                    for _, row in df.iterrows():
                        cursor.execute("""
                            INSERT INTO Loans (
                                "Serial Number","LENDER'S NAME ", "Loan Amount", "Interest", "Period", "Emi Amount" ,"Emi Due Date","Installment No","Principle Remaining","Principle Outstanding","Date","Classification"
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            row['S. No.'],
                            row["LENDER'S NAME "],
                            row['Loan amount'],
                            row['Interest'],
                            row['Period (months)'],
                            row['EMI AMOUNT'],
                            row['EMI DUE DATE'],
                            row['Installment No'],
                            row['Principle Remaining'],
                            row['Principle Outstanding'],
                            row['Date'],
                            row['Classification']
                        ))
                if detected_classification == "OD":
                    for _, row in df.iterrows():
                        cursor.execute("""
                            INSERT INTO Loans (
                                "Serial Number", "Interest", "LENDER'S NAME ","Utilisation Amount","Balance Unused","Classification"
                            ) VALUES (?, ?, ?, ?, ?, ?)
                        """, (
                            row['S. No.'],
                            row['Interest'],
                            row["LENDER'S NAME "],
                            row['Utilisation Amount'],
                            row['Balance Unused'],
                            row['Classification']
                        ))
                if detected_classification == "Secured":
                    for _, row in df.iterrows():
                        cursor.execute("""
                            INSERT INTO Loans (
                                "Serial Number","LENDER'S NAME ", "Loan Amount", "Interest", "Period", "Emi Amount" ,"Installment No","Principle Remaining","Principle Outstanding","Date","Classification"
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            row['S. No.'],
                            row["LENDER'S NAME "],
                            row['Loan amount'],
                            row['Interest'],
                            row['Period (months)'],
                            row['EMI AMOUNT'],
                            row['Installment No'],
                            row['Principle Remaining'],
                            row['Principle Outstanding'],
                            row['Date'],
                            row['Classification']
                        ))

            if flag == 3:
                for _, row in df.iterrows():
                    cursor.execute("""
                        INSERT INTO Sales (
                            "Serial Number", 
                            "Item Head", 
                            "Value", 
                            "Date", 
                            "Classification"
                        ) VALUES (?, ?, ?, ?, ?)
                    """, (
                        row['S.No'],
                        row['Project / Item head'],
                        row['Value'],
                        row['Date'],
                        row['Classification']
                    ))

            conn.commit()
            conn.close()
            st.success("Data saved to SQLite database successfully!")

    except Exception as e:
        st.warning("Something went wrong")
        st.text(f"Error: {e}")
